import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename


def merge_contract_and_backlog():
    # Use tkinter to open file dialogs for selecting files
    Tk().withdraw()  # Hide the root Tk window

    # Prompt the user to select the contract file
    contract_file = askopenfilename(title="Select the Contract File for Neotech", filetypes=[("Excel files", "*.xlsx")])
    if not contract_file:
        print("No contract file selected.")
        return

    # Prompt the user to select the backlog file
    backlog_file = askopenfilename(title="Select the Backlog File for Neotech", filetypes=[("Excel files", "*.xlsx")])
    if not backlog_file:
        print("No backlog file selected.")
        return

    # Load the specific sheet 'Dupes Removed' from the contract file
    contract_df = pd.read_excel(contract_file, sheet_name='Dupes Removed')

    # Strip any leading/trailing whitespace from column names
    contract_df.columns = contract_df.columns.str.strip()

    # Print columns to debug
    print("Columns in the contract file:", contract_df.columns.tolist())

    # Ensure PARTNUM and Backlog CPN are strings and case-insensitive
    if 'PARTNUM' in contract_df.columns:
        contract_df['PARTNUM'] = contract_df['PARTNUM'].astype(str).str.lower()
    else:
        print("Column 'PARTNUM' not found in the contract file.")
        return

    backlog_df = pd.read_excel(backlog_file)
    backlog_df['Backlog CPN'] = backlog_df['Backlog CPN'].astype(str).str.lower()

    # Create a new column 'Contract Price' with NaN values
    backlog_df['Contract Price'] = None

    # Compare and pull data into 'Contract Price' column
    for i, row in backlog_df.iterrows():
        match = contract_df[contract_df['PARTNUM'] == row['Backlog CPN']]
        if not match.empty:
            backlog_df.at[i, 'Contract Price'] = match.iloc[0]['BASEUNITPRICE']  # Use the correct column name here

    # Prompt the user to select a location to save the output file
    output_file = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not output_file:
        print("No output file location selected.")
        return

    # Save the DataFrame to the selected Excel file
    backlog_df.to_excel(output_file, index=False)

    # Load the Excel file using openpyxl to add the formula
    wb = load_workbook(output_file)
    ws = wb.active

    # Find the column index of 'Contract Price'
    contract_price_col = ws.max_column

    # Yellow fill for headers
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Highlight the header of the 'Contract Price' column in yellow
    ws.cell(row=1, column=contract_price_col).fill = yellow_fill

    # Add the 'Diff' column with the Excel formula and apply yellow fill
    diff_col = contract_price_col + 1
    ws.cell(row=1, column=diff_col, value="Diff").fill = yellow_fill

    # Find the 'Backlog Resale' and 'Qty Sched' columns
    backlog_resale_col = None
    qty_sched_col = None

    for col in range(1, contract_price_col):
        header_value = ws.cell(row=1, column=col).value
        if header_value == 'Backlog Resale':
            backlog_resale_col = col
        elif header_value == 'Qty Sched':
            qty_sched_col = col

    if backlog_resale_col is None:
        print("Column 'Backlog Resale' not found in the backlog file.")
        return

    if qty_sched_col is None:
        print("Column 'Qty Sched' not found in the backlog file.")
        return

    # Add the Excel formula to the 'Diff' column
    for row in range(2, ws.max_row + 1):
        backlog_resale_col_letter = ws.cell(row=1, column=backlog_resale_col).column_letter
        contract_price_col_letter = ws.cell(row=1, column=contract_price_col).column_letter
        ws.cell(row=row, column=diff_col).value = f"={backlog_resale_col_letter}{row}-{contract_price_col_letter}{row}"

    # Add the 'Ext Value' column with the Excel formula and apply yellow fill
    ext_value_col = diff_col + 1
    ws.cell(row=1, column=ext_value_col, value="Ext Value").fill = yellow_fill

    for row in range(2, ws.max_row + 1):
        diff_col_letter = ws.cell(row=1, column=diff_col).column_letter
        qty_sched_col_letter = ws.cell(row=1, column=qty_sched_col).column_letter
        ws.cell(row=row, column=ext_value_col).value = f"={diff_col_letter}{row}*{qty_sched_col_letter}{row}"

    # Wrap text in headers and enable filters
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True)
    ws.auto_filter.ref = ws.dimensions

    # Save the workbook
    wb.save(output_file)

    # Show a message box to inform the user that the process is complete
    messagebox.showinfo("Process Complete", "The Excel file has been updated and saved successfully.")

    print(f"Data merged and saved to {output_file}")
